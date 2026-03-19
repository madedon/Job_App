"""
Pipeline Master Builder
Merges all historical Job_Alerts files into a single authoritative tracker.
Saves to: applications/job_pipeline_master.xlsx

This is the single source of truth for the job application pipeline.

Usage:
    python tools/build_pipeline_master.py            # build/rebuild from scratch
    python tools/build_pipeline_master.py --update "Company" Field Value
    python tools/build_pipeline_master.py --add "Company" "Role"
"""

import sys, io, os, re, argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from pathlib import Path

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    except Exception:
        pass

# ─────────────────────────────────────────────
# PATHS — all relative to project root
# ─────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).parent.parent
TMP = PROJECT_ROOT / ".tmp"
OUTPUT_DIR = PROJECT_ROOT / "applications"
OUTPUT_PATH = OUTPUT_DIR / "job_pipeline_master.xlsx"

# ─────────────────────────────────────────────
# MASTER SCHEMA
# ─────────────────────────────────────────────

MASTER_COLUMNS = [
    "Processed Date", "Company", "Role", "Location", "Type",
    "Compensation", "Source", "Tier", "Match %", "Status",
    "Pre-Screen Decision", "Pre-Screen Reason",
    "CV Done", "Cover Letter Done", "Applied Date",
    "Interview Stage", "Application Link", "Job Posting #", "Notes",
]

STATUS_COLORS = {
    "Applied":               "C6EFCE",
    "Deferred - Next Week":  "FFEB9C",
    "Skipped":               "D9D9D9",
    "Expired":               "F4CCCC",
    "Interviewing":          "9FC5E8",
    "Offer":                 "00FF00",
    "Rejected":              "EA4335",
    "Ready to Apply":        "BDD7EE",
    "CV Generated":          "E2EFDA",
    "Auto-Skipped":          "EFEFEF",
}

STATUS_PRIORITY = {
    "Interviewing": 10, "Offer": 9, "Applied": 8,
    "Ready to Apply": 7, "CV Generated": 6,
    "Deferred - Next Week": 5, "Skipped": 3,
    "Expired": 2, "Auto-Skipped": 1, "Rejected": 0,
}

COLUMN_WIDTHS = {
    "Processed Date": 14, "Company": 22, "Role": 42, "Location": 18,
    "Type": 12, "Compensation": 16, "Source": 14, "Tier": 6,
    "Match %": 10, "Status": 22, "Pre-Screen Decision": 18,
    "Pre-Screen Reason": 30, "CV Done": 10, "Cover Letter Done": 16,
    "Applied Date": 14, "Interview Stage": 18, "Application Link": 30,
    "Job Posting #": 16, "Notes": 55,
}


def normalize_company(company_str: str) -> str:
    if not isinstance(company_str, str):
        return ""
    return re.sub(r'\s+', ' ', company_str.strip().lower())


def extract_role(position_str: str) -> str:
    if not isinstance(position_str, str):
        return "Unknown"
    s = position_str.strip()
    m = re.search(r'\d+%\s+match\s+(.+?)\s+role\b', s, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r'hiring for\s+[\"\u201c\u2018\'](.*?)[\"\u201d\u2019\']', s, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r'^(.+?)\s+at\s+\w', s)
    if m:
        role = m.group(1).strip()
        if len(role) < 80 and not role.endswith("just posted"):
            return role
    return s[:80] if len(s) > 80 else s


def write_excel(df: pd.DataFrame, path: Path):
    os.makedirs(path.parent, exist_ok=True)
    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Pipeline Master", index=False)
        ws = writer.sheets["Pipeline Master"]
        header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, col_name in enumerate(df.columns, 1):
            col_letter = ws.cell(row=1, column=i).column_letter
            width = COLUMN_WIDTHS.get(col_name, 15)
            ws.column_dimensions[col_letter].width = width
        status_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Status":
                status_col_idx = idx
                break
        if status_col_idx:
            for row in ws.iter_rows(min_row=2):
                status_val = row[status_col_idx - 1].value
                if status_val and status_val in STATUS_COLORS:
                    fill = PatternFill(start_color=STATUS_COLORS[status_val],
                                       end_color=STATUS_COLORS[status_val], fill_type="solid")
                    for cell in row:
                        cell.fill = fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 28
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 18
    print(f"\nSaved to: {path}")


def update_master_row(company: str, updates: dict, role_hint: str = "") -> bool:
    """Update one or more fields for a company entry in the master file."""
    if not OUTPUT_PATH.exists():
        print(f"ERROR: Master file not found at {OUTPUT_PATH}")
        return False

    df = pd.read_excel(str(OUTPUT_PATH), sheet_name="Pipeline Master")
    company_key = normalize_company(company)
    df["_key"] = df["Company"].apply(normalize_company)
    matches = df[df["_key"] == company_key]

    if matches.empty:
        matches = df[df["_key"].str.contains(company_key[:10], na=False)]

    if matches.empty:
        new_row = {col: "" for col in MASTER_COLUMNS}
        new_row["Company"] = company
        new_row["Processed Date"] = datetime.now().strftime("%Y-%m-%d")
        if role_hint:
            new_row["Role"] = role_hint
        for field, value in updates.items():
            if field in new_row:
                new_row[field] = value
        df = df.drop(columns=["_key"])
        new_df = pd.DataFrame([new_row], columns=MASTER_COLUMNS)
        df = pd.concat([df, new_df], ignore_index=True)
        write_excel(df, OUTPUT_PATH)
        print(f"  Added new entry: {company}")
        return True

    if len(matches) > 1 and role_hint:
        role_key = role_hint.lower()
        role_matches = matches[matches["Role"].str.lower().str.contains(role_key[:20], na=False)]
        if not role_matches.empty:
            matches = role_matches

    idx = matches.index[0]
    df = df.drop(columns=["_key"])

    changed_fields = []
    for field, value in updates.items():
        if field in df.columns:
            df.at[idx, field] = value
            changed_fields.append(f"{field}={value}")
    write_excel(df, OUTPUT_PATH)
    print(f"  Updated {company}: {', '.join(changed_fields)}")
    return True


def add_new_position(company: str, role: str, processed_date: str = "", **kwargs) -> bool:
    """Add a brand-new position to the master."""
    if not OUTPUT_PATH.exists():
        print(f"Master file not found -- creating new one")
        df = pd.DataFrame(columns=MASTER_COLUMNS)
    else:
        df = pd.read_excel(str(OUTPUT_PATH), sheet_name="Pipeline Master")
        company_key = normalize_company(company)
        df["_key"] = df["Company"].apply(normalize_company)
        if (df["_key"] == company_key).any():
            print(f"  SKIP: {company} already in master")
            df = df.drop(columns=["_key"])
            return False
        df = df.drop(columns=["_key"])

    new_row = {col: "" for col in MASTER_COLUMNS}
    new_row["Processed Date"] = processed_date or datetime.now().strftime("%Y-%m-%d")
    new_row["Company"] = company
    new_row["Role"] = role
    for field, value in kwargs.items():
        col_match = field.replace("_", " ").title()
        if col_match in MASTER_COLUMNS:
            new_row[col_match] = value
        elif field in MASTER_COLUMNS:
            new_row[field] = value

    new_df = pd.DataFrame([new_row], columns=MASTER_COLUMNS)
    df = pd.concat([df, new_df], ignore_index=True)
    write_excel(df, OUTPUT_PATH)
    print(f"  Added: {company} -- {role}")
    return True


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Pipeline Master Builder")
    parser.add_argument("--update", nargs=3, metavar=("COMPANY", "FIELD", "VALUE"))
    parser.add_argument("--add", nargs=2, metavar=("COMPANY", "ROLE"))
    args = parser.parse_args()

    if args.update:
        company, field, value = args.update
        update_master_row(company, {field: value})
    elif args.add:
        company, role = args.add
        add_new_position(company, role)
    else:
        print("Usage:")
        print("  --update 'Company' Field Value")
        print("  --add 'Company' 'Role'")
