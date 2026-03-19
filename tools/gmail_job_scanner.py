"""
Gmail Job Alert Scanner
Scans Gmail for job alerts from the last N days, extracts relevant positions,
and creates an Excel file matching the weekly pipeline format.

Usage: python gmail_job_scanner.py [--days N]
Requires:
  - credentials.json (Google OAuth client credentials) in project root
  - First run will open browser for authorization
  - Token cached in token.json for subsequent runs
"""

import sys, io

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    except Exception:
        pass

import os
import re
import json
import base64
import argparse
from datetime import datetime, timedelta
from pathlib import Path

try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
except ImportError:
    print("ERROR: Google API libraries not installed.")
    print("Run: pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
    exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# Configuration — all paths relative to project root
PROJECT_ROOT = Path(__file__).parent.parent
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
CREDENTIALS_FILE = PROJECT_ROOT / "credentials.json"
TOKEN_FILE = PROJECT_ROOT / "token.json"
OUTPUT_DIR = PROJECT_ROOT / ".tmp"

# Keywords matching target profile
TITLE_KEYWORDS_STRONG = [
    "director", "senior director", "sr director", "sr. director",
    "vp", "vice president", "head of",
    "chief", "c-level", "cto", "cio", "cdo",
    "product director",
]
TITLE_KEYWORDS_GOOD = [
    "product manager", "product owner", "sr product manager", "sr. product manager",
    "senior product manager", "principal product manager",
    "ai lead", "ai manager", "ml lead", "aiml", "ai engineering",
    "data scientist", "data science manager", "ai solutions",
    "technical program manager", "program manager", "program director",
    "principal program manager", "sr program manager",
    "engineering manager", "senior engineering manager", "software development manager",
    "data science", "data engineering", "analytics director", "analytics manager",
    "senior manager", "principal", "lead", "associate director",
]
DOMAIN_KEYWORDS = [
    "ai", "artificial intelligence", "machine learning", "ml", "genai", "llm",
    "deep learning", "nlp", "natural language processing", "computer vision",
    "model lifecycle", "data pipelines", "prompt", "rag", "agentic",
    "telecom", "telecommunications", "5g", "network", "ran", "wireless",
    "cloud", "aws", "azure", "gcp", "google cloud", "data center",
    "infrastructure", "devops", "platform",
    "automation", "workflow automation", "process optimization",
    "digital transformation", "operations",
    "product management", "program management", "product development",
    "product delivery", "service delivery", "professional services",
    "consulting", "advisory", "strategy", "transformation",
    "engagement manager", "solutions architect",
    "data science", "data engineering", "analytics", "business intelligence",
    "agile", "scrum", "lean", "six sigma",
]

JOB_ALERT_SENDERS = [
    "jobs-noreply@linkedin.com",
    "jobalerts-noreply@linkedin.com",
    "notifications-noreply@linkedin.com",
    "invitations@linkedin.com",
    "messages-noreply@linkedin.com",
    "@indeed.com", "@indeedmail.com", "@glassdoor.com",
    "@jobot.com", "@jobright.ai",
    "@ziprecruiter.com", "@ziprecruiter.email",
    "@google.com", "@lever.co", "@greenhouse.io",
    "@myworkdayjobs.com", "@smartrecruiters.com",
    "@icims.com", "@dice.com", "@hired.com",
    "@wellfound.com", "@talent.com",
    "@careerbuilder.com", "@simplyhired.com",
    "@otta.com", "@builtin.com",
]


def authenticate():
    """Authenticate with Gmail API using OAuth 2.0."""
    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(f"  Token refresh failed ({e}). Re-authenticating via browser...")
                creds = None
        if not creds or not creds.valid:
            if not CREDENTIALS_FILE.exists():
                print(f"ERROR: {CREDENTIALS_FILE} not found.")
                print("\nTo set up Gmail access:")
                print("1. Go to https://console.cloud.google.com/")
                print("2. Create/select a project")
                print("3. Enable Gmail API")
                print("4. Create OAuth 2.0 credentials (Desktop app)")
                print(f"5. Download JSON and save as: {CREDENTIALS_FILE}")
                return None

            flow = InstalledAppFlow.from_client_secrets_file(
                str(CREDENTIALS_FILE), SCOPES
            )
            creds = flow.run_local_server(port=0)

        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())

    return creds


def search_job_emails(service, days_back=7):
    """Search Gmail for job-related emails from the last N days."""
    after_date = (datetime.now() - timedelta(days=days_back)).strftime("%Y/%m/%d")

    sender_queries = " OR ".join([f"from:{s}" for s in JOB_ALERT_SENDERS])
    subject_queries = " OR ".join([
        "subject:\"job alert\"", "subject:\"new job\"", "subject:\"jobs for you\"",
        "subject:\"new opportunities\"", "subject:\"recommended jobs\"",
        "subject:\"daily job\"", "subject:\"weekly job\"", "subject:hiring",
        "subject:\"is hiring\"", "subject:\"job match\"", "subject:\"apply now\"",
        "subject:\"open position\"", "subject:\"new role\"",
        "subject:\"career opportunity\"", "subject:director",
        "subject:\"vice president\"", "subject:\"VP \"", "subject:\"head of\"",
        "subject:\"job recommendation\"", "subject:\"your job search\"",
        "subject:\"we found\"", "subject:\"based on your profile\"",
        "subject:\"application\"", "subject:\"invitation to apply\"",
    ])
    body_queries = " OR ".join([
        "\"job alert\"", "\"is hiring\"", "\"apply now\"", "\"view job\"", "\"see job\"",
    ])

    query = f"after:{after_date} ({sender_queries} OR {subject_queries} OR {body_queries})"
    print(f"  Search window: last {days_back} days (after {after_date})")

    messages = []
    page_token = None
    while True:
        result = service.users().messages().list(
            userId="me", q=query, maxResults=100, pageToken=page_token
        ).execute()
        if "messages" in result:
            messages.extend(result["messages"])
        page_token = result.get("nextPageToken")
        if not page_token:
            break

    print(f"  Found {len(messages)} job-related emails")
    return messages


def extract_job_details(service, message_id):
    """Extract job details from an email message."""
    msg = service.users().messages().get(userId="me", id=message_id, format="full").execute()
    headers = {h["name"].lower(): h["value"] for h in msg["payload"]["headers"]}
    subject = headers.get("subject", "")
    sender = headers.get("from", "")
    date_str = headers.get("date", "")

    source = "Unknown"
    sender_lower = sender.lower()
    for name, pattern in [("LinkedIn", "linkedin"), ("Indeed", "indeed"),
                          ("Glassdoor", "glassdoor"), ("Jobot", "jobot"),
                          ("Jobright", "jobright"), ("ZipRecruiter", "ziprecruiter"),
                          ("Google Careers", "google"), ("Dice", "dice"),
                          ("Hired", "hired"), ("Wellfound", "wellfound"),
                          ("BuiltIn", "builtin")]:
        if pattern in sender_lower:
            source = name
            break

    body = _extract_body(msg["payload"])
    links = re.findall(r'https?://[^\s<>"\']+', body)
    apply_links = [l for l in links if any(kw in l.lower() for kw in
        ["apply", "job", "career", "position", "view", "opening"])]

    company = _extract_company(subject, body, source)
    location = _extract_location(body)
    salary = _extract_salary(body)
    posting_id = _extract_posting_id(links, source)
    text = (subject + " " + body).lower()
    fit_score, tier, notes = _calculate_fit(text, subject)

    return {
        "tier": tier, "company": company, "position": subject[:120],
        "location": location, "salary_range": salary, "posting_id": posting_id,
        "application_link": apply_links[0] if apply_links else (links[0] if links else ""),
        "source": source, "status": "Not Applied", "fit_score": fit_score,
        "fit_stars": _score_to_stars(fit_score), "notes": notes,
        "date": date_str, "body_preview": body[:1000],
    }


def _extract_body(payload):
    body = ""
    if "parts" in payload:
        for part in payload["parts"]:
            if part["mimeType"] == "text/plain" and "data" in part.get("body", {}):
                body = base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="replace")
                break
            elif part["mimeType"] == "text/html" and "data" in part.get("body", {}):
                body = base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="replace")
            elif "parts" in part:
                body = _extract_body(part)
                if body:
                    break
    elif "body" in payload and "data" in payload["body"]:
        body = base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="replace")
    return body


def _extract_company(subject, body, source):
    patterns = [
        r'(?:at|from)\s+([A-Z][A-Za-z0-9\s&.]+?)(?:\s+and|\s+in|\s+for|,|\.|$)',
        r'([A-Z][A-Za-z0-9\s&.]+?)\s+(?:is hiring|just posted)',
        r'([A-Z][A-Za-z0-9\s&.]+?)\s+(?:and \d+ more)',
    ]
    for pattern in patterns:
        match = re.search(pattern, subject)
        if match:
            company = match.group(1).strip()
            if 2 < len(company) < 50:
                return company
    return ""


def _extract_location(body):
    patterns = [
        r'(?:Location|location|Located in|located in)[:\s]+([A-Za-z\s,]+(?:TX|CA|NY|WA|IL|Remote)[A-Za-z\s,]*)',
        r'((?:Remote|Hybrid|On-site)\s*[-\u2013]\s*[A-Za-z\s,]+)',
        r'([A-Za-z]+,\s*TX\b)',
        r'(Remote)',
    ]
    for pattern in patterns:
        match = re.search(pattern, body[:2000])
        if match:
            return match.group(1).strip()[:60]
    return ""


def _extract_salary(body):
    patterns = [
        r'\$[\d,]+[kK]?\s*[-\u2013]\s*\$[\d,]+[kK]?',
        r'\$[\d,]+\s*[-\u2013]\s*\$[\d,]+',
        r'(?:salary|compensation|pay)[:\s]*\$[\d,]+',
    ]
    for pattern in patterns:
        match = re.search(pattern, body[:3000])
        if match:
            return match.group(0).strip()[:40]
    return ""


def _extract_posting_id(links, source):
    ids = []
    for link in links[:10]:
        if "linkedin.com" in link:
            match = re.search(r'/view/(\d+)', link)
            if match:
                ids.append(f"LI:{match.group(1)}")
        elif "indeed.com" in link:
            match = re.search(r'jk=([a-f0-9]+)', link)
            if match:
                ids.append(f"IN:{match.group(1)}")
        elif "glassdoor.com" in link:
            match = re.search(r'jobListingId=(\d+)', link)
            if match:
                ids.append(f"GD:{match.group(1)}")
    return " | ".join(ids[:3]) if ids else ""


def _calculate_fit(text, subject):
    score = 0
    matched = []
    subject_lower = subject.lower()
    for kw in TITLE_KEYWORDS_STRONG:
        if kw in subject_lower:
            score += 3
            matched.append(f"Title: {kw}")
    for kw in TITLE_KEYWORDS_GOOD:
        if kw in subject_lower:
            score += 2
            matched.append(f"Title: {kw}")
    for kw in DOMAIN_KEYWORDS:
        if kw in text:
            score += 1
            matched.append(kw)
    if score >= 7:
        tier = "T1"
    elif score >= 5:
        tier = "T2"
    elif score >= 3:
        tier = "T3"
    else:
        tier = "T4"
    unique_matched = list(dict.fromkeys(matched))[:6]
    notes = ", ".join(unique_matched) if unique_matched else "Low keyword match"
    return score, tier, notes


def _score_to_stars(score):
    if score >= 7: return "\u2605\u2605\u2605\u2605\u2605"
    elif score >= 5: return "\u2605\u2605\u2605\u2605"
    elif score >= 3: return "\u2605\u2605\u2605"
    elif score >= 2: return "\u2605\u2605"
    else: return "\u2605"


def deduplicate(jobs):
    seen = {}
    unique = []
    for job in jobs:
        key = (job["company"].lower().strip(), re.sub(r'[^a-z]', '', job["position"].lower())[:30])
        posting_key = job["posting_id"]
        if posting_key and posting_key in seen:
            existing = seen[posting_key]
            if job["source"] not in existing["source"]:
                existing["source"] += f" + {job['source']}"
            continue
        if key in seen and not posting_key:
            continue
        seen[key] = job
        if posting_key:
            seen[posting_key] = job
        unique.append(job)
    return unique


def create_excel(jobs, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Alerts"
    headers = ["Tier", "Company", "Position", "Location", "Salary Range",
               "Job Posting #", "Application Link", "Source", "Status",
               "Fit Score", "Notes", "CV Approved", "CL Approved"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")
    tier_fills = {
        "T1": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "T2": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "T3": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "T4": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    tier_order = {"T1": 0, "T2": 1, "T3": 2, "T4": 3}
    jobs.sort(key=lambda x: (tier_order.get(x["tier"], 9), -x["fit_score"]))
    for row_idx, job in enumerate(jobs, 2):
        data = [job["tier"], job["company"], job["position"], job["location"],
                job["salary_range"], job["posting_id"], job["application_link"],
                job["source"], job["status"], job["fit_stars"], job["notes"], "", ""]
        tier_fill = tier_fills.get(job["tier"], tier_fills["T4"])
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value)[:500])
            cell.border = thin_border
            cell.fill = tier_fill
            if col_idx in (1, 4, 8, 9, 10, 12, 13):
                cell.alignment = center
    widths = {"A": 6, "B": 25, "C": 55, "D": 25, "E": 18, "F": 25,
              "G": 60, "H": 15, "I": 14, "J": 12, "K": 50, "L": 14, "M": 14}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    print(f"  Excel saved: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Scan Gmail for job alerts")
    parser.add_argument("--days", type=int, default=7, help="Days back to scan (default: 7)")
    args = parser.parse_args()

    print("=" * 60)
    print("GMAIL JOB ALERT SCANNER")
    print(f"Scanning last {args.days} days")
    print("=" * 60)

    print("\n[1] Authenticating with Gmail...")
    creds = authenticate()
    if not creds:
        return

    service = build("gmail", "v1", credentials=creds)

    print("\n[2] Searching for job alert emails...")
    messages = search_job_emails(service, days_back=args.days)
    if not messages:
        print("  No job alert emails found.")
        return

    limit = min(len(messages), 150)
    print(f"\n[3] Extracting details from {limit} emails...")
    jobs = []
    for i, msg in enumerate(messages[:limit]):
        try:
            job = extract_job_details(service, msg["id"])
            if job["fit_score"] > 0:
                jobs.append(job)
            if (i + 1) % 25 == 0:
                print(f"  Processed {i + 1}/{limit} emails...")
        except Exception as e:
            print(f"  Error processing message {msg['id']}: {e}")
    print(f"  Extracted {len(jobs)} relevant listings")

    print("\n[4] Deduplicating...")
    jobs = deduplicate(jobs)
    print(f"  After dedup: {len(jobs)} unique listings")

    print("\n[5] Creating Excel...")
    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d")
    output_path = OUTPUT_DIR / f"Job_Alerts_{timestamp}.xlsx"
    create_excel(jobs, output_path)

    tier_counts = {}
    for j in jobs:
        tier_counts[j["tier"]] = tier_counts.get(j["tier"], 0) + 1

    print(f"\n{'=' * 60}")
    print("SCAN COMPLETE")
    print(f"Total emails scanned: {len(messages)}")
    print(f"Relevant listings: {len(jobs)}")
    for tier in ["T1", "T2", "T3", "T4"]:
        if tier in tier_counts:
            print(f"  {tier}: {tier_counts[tier]} positions")
    print(f"Output: {output_path}")
    print(f"{'=' * 60}")

    print("\nTOP MATCHES:")
    for i, job in enumerate(jobs[:10], 1):
        print(f"  {i}. [{job['tier']}] {job['fit_stars']} [{job['source']}] {job['position'][:70]}")
        if job["company"]:
            print(f"     Company: {job['company']}")


if __name__ == "__main__":
    main()
